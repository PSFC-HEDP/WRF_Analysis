from __future__ import annotations

import argparse
import csv
import os
import re
from collections import OrderedDict
from math import sqrt, pi, nan, inf
from typing import Any, Optional

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from matplotlib.ticker import ScalarFormatter
from numpy.typing import NDArray
from scipy import optimize

from load_info_from_nif_database import normalize_shot_number
from src.calculate_rhoR import perform_hohlraum_correction, calculate_rhoR, Layer, Peak, np_Peak, Quantity, np_Quantity

# matplotlib.use("qtagg")
np.seterr(all="raise", under="ignore")
plt.rcParams["font.size"] = 12

# key values that the program needs to know
ROOT = 'data'
σWRF = .159
δσWRF = .014

# define an analysis type that combines all the information about one WRF spectrum
Analysis = tuple[str, str, str, str, str, bool, bool, Peak, Quantity, Peak, Quantity, NDArray[float]]
np_Analysis = np.dtype([
	("shot_day", np.str_, 11), ("shot_number", np.str_, 7),
	("line_of_site", np.str_, 7), ("position", np.str_, 2),
	("tag", np.str_, 16), ("overlapd", bool), ("clipd", bool),
	("peak", np_Peak), ("rhoR", np_Quantity),
	("compression", np_Peak), ("compression_rhoR", np_Quantity),
	("spectrum", object)])
# define an analysis type that combines all the information we get from yield ratios
SecondaryAnalysis = tuple[Quantity, Quantity]
np_SecondaryAnalysis = np.dtype([("rhoR", np_Quantity), ("temperature", np_Quantity)])


def make_plots_from_analysis(folders: list[str], show_plots: bool, command_line_options: dict[str, Any]):
	""" take the analysis .csv files created by AnalyzeCR39 in a given series of folders, and
	    generate a bunch of plots and tables summarizing the information therein.
	    :param folders: a list of subdirectories in data/ to search for analysis results
	    :param show_plots: whether to show the plots as they’re generated in addition to saving them to disk
	    :param command_line_options: additional values specified in the original command
	"""
	analyses = []
	for i, folder in enumerate(folders):
		if i > 0 and len(folder) < 7: # allow user to only specify the last few digits when most of the foldername is the same
			folders[i] = folders[i-1][:-len(folder)] + folder

	# first, load the analyzed data
	for i, folder_name in enumerate(folders): # for each specified folder
		folder = os.path.join(ROOT, folder_name)
		if not os.path.isdir(folder):
			print(f"No such folder: `{folder}`")
			return

		for subfolder, _, _ in os.walk(folder): # and any subfolders inside it
			for filename in os.listdir(subfolder): # scan all files inside that folder
				if re.fullmatch(r'(A|N|Om?)\d{6}-?\d{3}.csv', filename): # if it is a shot summary file
					analyses += read_shot_summary_file(os.path.join(subfolder, filename))

				elif re.fullmatch(r'.*ANALYSIS.*\.csv', filename): # if it is an analysis file
					try:
						analyses.append(read_analysis_file(
							folder, os.path.join(subfolder, filename), show_plots, command_line_options))
					except (HohlraumFileError, MetadataNotFoundError) as e:
						print(e)
						return

	if len(analyses) == 0:
		print("no datum were found.")
		return

	analyses = np.array(analyses, dtype=np_Analysis)

	# compose labels of the appropriate specificity
	multiple_days = not np.all(analyses["shot_day"] == analyses["shot_day"][0])
	multiple_shots = multiple_days or \
		not np.all(analyses["shot_number"] == analyses["shot_number"][0])
	shot_labels = []
	labels = []
	for analysis in analyses:
		shot_label, analysis_label = assign_label(analysis, multiple_days, multiple_shots)
		shot_labels.append(shot_label)
		labels.append(analysis_label)

	shot_labels = np.array(shot_labels)
	labels = np.array(labels)

	# convert sigmas to widths
	widths = np.empty(len(analyses), dtype=np_Quantity)
	widths["lower_err"] = analyses["peak"]["sigma"]["lower_err"]*2.355e3
	widths["value"] = analyses["peak"]["sigma"]["value"]*2.355e3
	widths["upper_err"] = analyses["peak"]["sigma"]["upper_err"]*2.355e3
	# convert sigmas to temperatures
	temperatures = np.empty(len(analyses), dtype=np_Quantity)
	temperatures["value"] = (np.sqrt(
		np.maximum(0, analyses["peak"]["sigma"]["value"]**2 - σWRF**2))/76.68115805e-3)**2
	temperatures["lower_err"] = np.sqrt(
		(2*analyses["peak"]["sigma"]["value"]*analyses["peak"]["sigma"]["lower_err"])**2 +
		(2*σWRF*δσWRF)**2
	)/76.68115805e-3**2
	temperatures["upper_err"] = temperatures["lower_err"]

	base_directory = os.path.join(ROOT, folders[0])

	# load any results we got from Patrick's secondary analysis
	try:
		secondary_stuff = np.atleast_2d(np.loadtxt(
			os.path.join(base_directory, f'secondary_stuff.txt'), dtype=float))
	except IOError:
		secondary_stuff = None # but it's okey if there is none
	if secondary_stuff is None or np.all(np.isnan(secondary_stuff)):
		secondary_stuff = np.full((len(analyses), 6), nan, dtype=float)
		np.savetxt(os.path.join(base_directory, f'secondary_stuff.txt'), secondary_stuff) # type: ignore
	secondary_analyses = np.empty(len(analyses), dtype=np_SecondaryAnalysis)
	for i, row in enumerate(secondary_stuff):
		secondary_analyses[i] = ((row[0], row[1], row[2]), (row[3], row[4], row[5]))
	assert len(secondary_analyses) == len(analyses), "This secondary analysis file has the rong number of entries"

	# save the spectra to a csv file
	for label, analysis in zip(labels, analyses):
		sanitized_label = re.sub(r"[:\s]", "_", label)
		filename = os.path.join(base_directory, f"spectrum_{sanitized_label}.csv")
		np.savetxt(filename, analysis["spectrum"],
		           header="Energy after passing through hohlraum (MeV),Spectrum (MeV^-1),Uncertainty (MeV^-1)",
		           delimiter=",", comments="")

	# save the spectra in a spreadsheet
	workbook = openpyxl.Workbook()
	worksheet = workbook.active
	for i, (label, analysis) in enumerate(zip(labels, analyses)):
		worksheet.cell(1, 4*i + 1).value = label.replace("\n", " ")
		worksheet.merge_cells(start_row=1, end_row=1, start_column=4*i + 1, end_column=4*i + 3)
		worksheet.cell(2, 4*i + 1).value = "Energy (MeV)"
		worksheet.cell(2, 4*i + 2).value = "Spectrum (MeV^-1)"
		worksheet.cell(2, 4*i + 3).value = "Error bar (MeV^-1)"
		for j in range(analysis["spectrum"].shape[0]):
			for k in range(analysis["spectrum"].shape[1]):
				worksheet.cell(3 + j, 4*i + 1 + k).value = analysis["spectrum"][j, k]
	workbook.save(os.path.join(base_directory, "WRF spectra.xlsx"))

	# save the condensed results in a spreadsheet for each folder
	workbooks = {}
	for item, secondary_stuff in zip(analyses, secondary_analyses):
		if not item["shot_day"].startswith("N"):
			continue  # only NIF shots do this part
		filename = f"{item['shot_day']}-{item['shot_number']}-999 {item['line_of_site']} WRF report.xlsx"
		if filename not in workbooks:
			workbooks[filename] = openpyxl.load_workbook("templates/report.xlsx")
		worksheet = workbooks[filename].active
		worksheet.cell(2, 2).value = f"{item['shot_day']}-{item['shot_number']}-999"
		worksheet.cell(3, 2).value = f"0{item['line_of_site']}"
		header_row = {"1": 10, "2": 20, "3": 40, "4": 30}[item["position"]]
		report_quantities = {
			1: item["peak"]["yield"], 2: item["compression"]["yield"],
			3: item["rhoR"], 4: item["compression_rhoR"]}
		for row, quantity in report_quantities.items():
			worksheet.cell(header_row + row, 2).value = quantity["value"]
			if worksheet.cell(header_row + row, 7).value == "perc":
				if quantity["value"] != 0:
					worksheet.cell(header_row + row, 6).value = quantity["upper_err"]/quantity["value"]
			else:
				worksheet.cell(header_row + row, 6).value = quantity["upper_err"]
	for filename, workbook in workbooks.items():
		workbook.save(os.path.join(base_directory, filename))

	# print out a table, and also save the condensed results in a csv file
	print()
	with open(os.path.join(base_directory, 'wrf_analysis.csv'), 'w') as f:
		print("|  WRF              |  Yield              |Mean energy (MeV)| ρR (mg/cm^2)  |")
		print("|-------------------|---------------------|----------------|----------------|")
		f.write(
			"WRF, Yield, Yield unc., Mean energy (MeV), Mean energy unc. (MeV), "
			"Temperature (keV), Temperature unc. (keV), Width (keV), Width unc. (keV), "
			"Rho-R (mg/cm^2), Rho-R unc. (mg/cm^2), "
			"Compres. yield, Compres. yield unc., Compres. mean (MeV), Compres. mean unc. (MeV), "
			"Compres. rho-R (mg/cm^2), Compres. rho-R unc. (mg/cm^2)"
			"\n")

		last_shot_label = None
		for shot_label, label, item, width, temperature in zip(shot_labels, labels, analyses, widths, temperatures):
			label = label.replace('\n', ' ')
			yeeld, mean, _ = item["peak"]
			compression_yield, compression_mean, compression_sigma = item["compression"]
			rhoR, compression_rhoR = item["rhoR"], item["compression_rhoR"]

			if last_shot_label is not None and last_shot_label != shot_label:
				print("")
			print("|  {:s}  |  {:#.2g} ± {:#.2g}  |  {:5.2f} ± {:4.2f}  |  {:5.1f} ± {:4.1f}  |".format(
				label[-15:],
				yeeld["value"], (yeeld["lower_err"] + yeeld["upper_err"])/2,
				mean["value"], (mean["lower_err"] + mean["upper_err"])/2,
				rhoR["value"], (rhoR["lower_err"] + rhoR["upper_err"])/2))

			f.write("{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{}\n".format(
				label,
				yeeld["value"], (yeeld["lower_err"] + yeeld["upper_err"])/2,
				mean["value"], (mean["lower_err"] + mean["upper_err"])/2,
				temperature["value"], (temperature["lower_err"] + temperature["upper_err"])/2,
				width["value"], (width["lower_err"] + width["upper_err"])/2,
				rhoR["value"], (rhoR["lower_err"] + rhoR["upper_err"])/2,
				compression_yield["value"], compression_yield["upper_err"],
				compression_mean["value"], compression_mean["upper_err"],
				compression_rhoR["value"], compression_rhoR["upper_err"]))

			last_shot_label = shot_label
	print()

	# make the error bars asymmetrick if there are issues with the data
	if np.any(analyses["overlapd"]):
		decreased_data = [analyses["peak"]["yield"], analyses["compression"]["yield"],
		                  secondary_analyses["rhoR"], secondary_analyses["temperature"]]
		distorted_data = [analyses["peak"]["sigma"], widths, temperatures]
		for data in decreased_data:
			data["upper_err"][analyses["overlapd"]] = 2e20
		for data in distorted_data:
			data["value"][analyses["overlapd"]] = nan
	if np.any(analyses["clipd"]):
		decreased_data = [analyses["peak"]["yield"], analyses["compression"]["yield"],
		                  analyses["rhoR"], analyses["compression_rhoR"],
		                  secondary_analyses["rhoR"], secondary_analyses["temperature"]]
		increased_data = [analyses["peak"]["mean"]]
		for data in decreased_data:
			data["upper_err"][analyses["clipd"]] = 2e20
		for data in increased_data:
			data["lower_err"][analyses["clipd"]] = 2e20

	# choose label spacing based on number of shots
	if len(analyses) <= 6:
		rotation = 0
		alignment = 'center'
		spacing = 1
	else:
		rotation = 45
		alignment = 'right'
		if len(analyses) <= 15 or len(np.unique(analyses["shot_day"])) > 1:
			spacing = 0.55
		else:
			spacing = 0.40

	# create the comparison plots
	if np.any(np.isfinite(analyses["compression"]["yield"]["value"])):
		yield_descriptor = "Shock yield"
		mean_descriptor = "Shock peak energy"
		rhoR_descriptor = "Shock ρR"
	else:
		yield_descriptor = "Yield"
		mean_descriptor = "Mean energy"
		rhoR_descriptor = "Total ρR"
	for label, filetag, data in [
		(f"{yield_descriptor}", 'yield', analyses["peak"]["yield"]),
		(f"{mean_descriptor} (MeV)", 'mean', analyses["peak"]["mean"]),
		(f"{rhoR_descriptor} (mg/cm^2)", 'rhoR_total', analyses["rhoR"]),
		("Compression yield", 'yield_compression', analyses["compression"]["yield"]),
		("Compression ρR (mg/cm^2)", 'rhoR_compression', analyses["compression_rhoR"]),
		("Fuel ρR (mg/cm^2)", 'rhoR_fuel', secondary_analyses["rhoR"]),
		("Electron temperature (keV)", 'temperature_electron', secondary_analyses["temperature"]),
		("Width (keV)", 'width', widths),
		("Ion temperature (keV)", 'temperature_ion', temperatures),
	]:
		if data.size == 0 or np.all(np.isnan(data["value"])): # skip if there's noting here
			continue

		plt.figure(figsize=(1.5+len(data)*spacing, 4.5)) # then plot it!
		plt.errorbar(x=np.arange(len(data)),
		             y=data["value"],
		             yerr=[data["lower_err"], data["upper_err"]],
		             fmt='.k', elinewidth=2, markersize=12)

		plt.xlim(-1/2, len(data) - 1/2)
		plt.xticks(ticks=np.arange(len(data)),
		           labels=labels,
		           rotation=rotation,
		           ha=alignment) # set the tick stuff
		plt.ylabel(label)
		plt.grid()

		if np.all(data["value"] == 0):
			continue
		max_value = np.max(data["value"]) # figure out the scale and limits
		min_value = np.min(data["value"], where=data["value"] != 0, initial=inf)
		tops = data["value"] + data["upper_err"]
		bottoms = data["value"] - data["lower_err"]
		points = np.stack([bottoms, data["value"], tops])
		measurable = data["value"] - np.minimum(data["lower_err"], data["upper_err"]) >= 0

		if np.any(measurable):
			valid = (points > -1e20) & (points < 1e20) & \
			        np.stack([measurable, np.full(measurable.shape, True), measurable])
		else:
			valid = (points > -1e20) & (points < 1e20)
		plot_top = np.max(points, where=valid, initial=-inf)
		plot_bottom = np.min(points, where=valid, initial=inf)
		if "MeV" in label and np.all(data["value"] < 14.7):
			plot_top = min(15, plot_top)

		if min_value > 0 and max_value/min_value > 10 and plot_bottom > 0:
			plt.yscale('log')
			plt.grid(axis='y', which='minor')
			rainge = plot_top/plot_bottom
			plt.ylim(plot_bottom/rainge**0.15, plot_top*rainge**0.15)
		else:
			rainge = plot_top - plot_bottom
			plt.ylim(max(0, plot_bottom - 0.15*rainge), plot_top + 0.15*rainge)

		# if filetag == "yield":
		# 	plt.ylim(-.1e11, 2e11)
		if filetag == "temperature_electron":
			plt.ylim(None, 4)
		plt_set_locators()

		plt.tight_layout()
		plt.savefig(os.path.join(base_directory, f'summary_{filetag}.png'), dpi=300)
		plt.savefig(os.path.join(base_directory, f'summary_{filetag}.eps'))

	# now create a line-of-site comparison plot
	compared_lines_of_site = np.unique(analyses["line_of_site"])
	if len(compared_lines_of_site) >= 2:
		compared_shots = sorted(set(shot_labels))
		los_rhoRs = np.empty((len(compared_shots), 2), dtype=np_Quantity)
		for i, shot in enumerate(compared_shots):
			for j, line_of_site in enumerate(compared_lines_of_site[:2]):
				here = (analyses["line_of_site"] == line_of_site) & (shot_labels == shot)
				if np.any(here) and not np.any(analyses["rhoR"]["lower_err"][here] == 0):
					los_rhoRs["value"][i, j] = np.average(
						analyses["rhoR"]["value"][here],
						weights=analyses["rhoR"]["lower_err"][here]**(-2))
					los_rhoRs["lower_err"][i, j] = np.min(analyses["rhoR"]["lower_err"][here])
					los_rhoRs["upper_err"][i, j] = np.min(analyses["rhoR"]["upper_err"][here])
				else:
					los_rhoRs[i, j] = (nan, nan, nan)

		if np.all(np.isfinite(los_rhoRs["value"])):
			plt.figure(figsize=(4.5, 4.5))
			plt_set_locators()
			plt.errorbar(y=los_rhoRs["value"][:, 0],
			             yerr=[los_rhoRs["lower_err"][:, 0],
			                   los_rhoRs["upper_err"][:, 0]],
			             x=los_rhoRs["value"][:, 1],
			             xerr=[los_rhoRs["lower_err"][:, 1],
			                   los_rhoRs["upper_err"][:, 1]],
			             fmt='.', color='#000000', markersize=12)
			plt.axline((0, 0), slope=1, color='k', linewidth=1)
			plt.ylabel(f"ρR on {compared_lines_of_site[0]}")
			plt.xlabel(f"ρR on {compared_lines_of_site[1]}")
			plt.axis('square')
			plt.xlim(0, np.max(los_rhoRs["value"])*1.4)
			plt.ylim(0, np.max(los_rhoRs["value"])*1.4)
			plt.grid()
			plt.tight_layout()
			plt.savefig(os.path.join(base_directory, f'summary_asymmetry.png'), dpi=300)
			plt.savefig(os.path.join(base_directory, f'summary_asymmetry.eps'))

	# show plots
	if show_plots:
		plt.show()
	plt.close('all')


def read_shot_summary_file(filepath: str) -> list[Analysis]:
	""" read a file that lists the key outputs from a single shot (as gets generated in the folders
	    that summarize entire campains) and put those outputs in a convenient format
	    :param filepath: the relative or absolute path to the summary file
	    :return: a list with an Analysis object describing each row
	"""
	shot_day, shot_number = os.path.splitext(os.path.basename(filepath))[0].split('-')

	analyses: list[Analysis] = []
	with open(filepath, encoding="utf8") as f:
		for row in f.readlines(): # read all the wrf summaries out of it
			values = re.sub(r"[|:±]", " ", row).split()
			line_of_site, position, yield_value, yield_error, \
				mean_value, mean_error, rhoR_value, rhoR_error = values

			analyses.append((
				shot_day, shot_number,
				line_of_site, position,
				"", False, False,
				(
					(float(yield_value), float(yield_error), float(yield_error)),
					(float(mean_value), float(mean_error), float(mean_error)),
					(nan, inf, inf),
				),
				(float(rhoR_value), float(rhoR_error), float(rhoR_error)),
				(
					(nan, inf, inf),
					(nan, inf, inf),
					(nan, inf, inf),
				),
				(nan, inf, inf),
				np.empty(0),
			))

	return analyses


def read_analysis_file(folder: str, filepath: str,
                       show_plots: bool, command_line_parameters: dict[str, Any]) -> Analysis:
	""" read an analysis file that came out of AnalyzeCR39 and pull out the key details in an Analysis struct
	    :param folder: the main folder to which this analysis file belongs
	    :param filepath: the relative or absolute path to the analysis file
	    :param show_plots: whether to show the plot that's generated in addition to saving it to disk
	    :param command_line_parameters: any ρR calculation information specified on the command line
	    :return: an Analysis object summarizing the analysis file
	    :raise HohlraumFileError: if hohlraum.txt is missing or invalid
	    :raise MetadataNotFoundError: if the analysis filename is missing some of the necessary metadata
	"""
	# read the filename for top-level metadata
	shot_day, shot_number, line_of_site, position, wrf_number = None, None, None, None, None
	tag = ''
	identifiers = re.split(r"[_/\\ ]", filepath.upper())
	for identifier in reversed(identifiers):
		if re.fullmatch(r'N\d{6}-?\d{3}-?999', identifier):
			shot_day, shot_number, _ = identifier.split('-')
		elif re.fullmatch(r'OM?2\d{5}', identifier):
			shot_day = identifier
		elif re.fullmatch(r'O?1?\d{5}', identifier):
			shot_number = identifier
		elif re.fullmatch(r'(DIM-?)?(0+-0+|0?90-(0?78|124|315))|TIM[1-6](-(4|8|12))?|(NDI-)?P2|NDI', identifier):
			line_of_site = identifier
		elif re.fullmatch(r'((POS|Pos)-?)?[1-4]|(4|8|12):00', identifier):
			position = identifier[-1]
		elif re.fullmatch(r'134\d{5}|G[0-2]\d{2}', identifier):
			wrf_number = identifier
		elif re.fullmatch(r'(LEFT|RIGHT|TOP|BOTTOM|MIDDLE|FULL)', identifier):
			tag = identifier.lower()

	if shot_number is None:
		raise MetadataNotFoundError(f"the filename is incomplete; there's no shot number in {identifiers}")
	if line_of_site is None:
		raise MetadataNotFoundError(f"the filename is incomplete; there's no line of sight in {identifiers}")
	if re.fullmatch(r"TIM[1-6]-(4|8|12)", line_of_site):  # extract the position from the line of site if relevant
		line_of_site, position = re.fullmatch(r"(TIM[1-6])-(4|8|12)", line_of_site).group(1, 2)
	elif re.fullmatch(r"\D*\d+-\d+", line_of_site): # standardize the DIM names if they're too long
		theta, phi = re.fullmatch(r"\D*(\d+)-(\d+)", line_of_site).group(1, 2)
		line_of_site = f"{int(theta):02d}-{int(phi):03d}"
	elif position is None:
		position = ""

	print(f"{wrf_number} – {shot_day}-{shot_number} {line_of_site}:{position} {tag}")

	# read thru the analysis file
	with open(filepath, newline='') as f:

		yeeld: Optional[Quantity] = None
		mean: Optional[Quantity] = None
		sigma: Optional[Quantity] = None
		gaussian_fit = True
		we_have_reachd_the_spectrum_part = False
		spectrum = []

		for row in csv.reader(f):
			if len(row) == 0:
				continue

			if not we_have_reachd_the_spectrum_part:
				if row[0] == 'Value (gaussian fit):':
					try:
						mean = (float(row[1]), 0., 0.)
						sigma = (float(row[2]), 0., 0.)
						yeeld = (float(row[3]), 0., 0.)
					except ValueError:
						gaussian_fit = False

				elif row[0] == 'Value (raw stats):' and not gaussian_fit:
					mean = (float(row[1]), 0., 0.)
					sigma = (float(row[2]), 0., 0.)
					yeeld = (float(row[3]), 0., 0.)

				elif row[0] == '    Random:' or row[0] == '    Systematic calib:':
					mean = (mean[0], mean[1] + float(row[1])**2, 0.)  # start by summing the errors as variances
					sigma = (sigma[0], sigma[1] + float(row[2])**2, 0.)
					yeeld = (yeeld[0], yeeld[1] + (float(row[3][:-1])*yeeld[0]/100)**2, 0.)

				elif row[0] == 'Energy 	 Yield/MeV 	Stat. Error':
					we_have_reachd_the_spectrum_part = True

			else:
				values = row[0].split()
				spectrum.append([float(v) for v in values])

	# then convert variances to sigmas (and also copy the lower error as the upper error)
	mean = (mean[0], sqrt(mean[1]), sqrt(mean[1]))
	sigma = (sigma[0], sqrt(sigma[1]), sqrt(sigma[1]))
	yeeld = (yeeld[0], sqrt(yeeld[1]), sqrt(yeeld[1]))

	# clean up the extracted spectrum
	spectrum = np.array(spectrum)
	spectrum = spectrum[spectrum[:, 2] != 0, :] # remove any points with sus error bars
	spectrum = spectrum[2:, :] # remove the two lowest bins because Fredrick’s program calculates them incorrectly

	# try to fit the compression peak
	if gaussian_fit:
		try:
			compression_fit = \
				fit_skew_gaussian(spectrum, 0, mean[0] - 2*sigma[0])
		except FitError:
			compression_fit = None
	else:
		compression_fit = None
	if compression_fit is not None:
		compression_yield, compression_mean, compression_sigma = compression_fit
		good_compression_fit = (compression_yield[0] > yeeld[0] and
		                        compression_mean[0] > spectrum[0][0] and
		                        compression_mean[0] < mean[0])
		if not good_compression_fit:
			compression_yield, compression_mean, compression_sigma = \
				(nan, inf, inf), (nan, inf, inf), (nan, inf, inf)
	else:
		compression_yield, compression_mean, compression_sigma = \
			(nan, inf, inf), (nan, inf, inf), (nan, inf, inf)
		good_compression_fit = False

	# load info from hohlraum.txt and the shot table
	parameters = load_rhoR_parameters(folder)
	parameters.update(command_line_parameters)
	any_hohlraum = any(parameters["hohlraum"].values())
	any_clipping_here = any(indicator in filepath for indicator in parameters["clipping"])
	any_overlap_here = any(indicator in filepath for indicator in parameters["overlap"])

	# plot its spectrum
	plt.figure(figsize=(10, 4))
	plt.grid()
	plt.axhline(0, color='k', linewidth=1)
	if gaussian_fit:
		x = np.linspace(0, 20, 1000)
		plt.plot(x, gaussian(
			x, yeeld[0], mean[0], sigma[0]),
		         color='#C00000')
		if compression_yield[0] > 0:
			plt.plot(x, skew_gaussian(
				x, compression_yield[0], compression_mean[0], compression_sigma[0]),
			         color='#C00000')
	plt.errorbar(x=spectrum[:, 0],
	             y=spectrum[:, 1],
	             yerr=spectrum[:, 2],
	             fmt='.', color='#000000', elinewidth=1, markersize=6)
	plt.axis([4, 18, min(0, np.min(spectrum[:, 1] + spectrum[:, 2])), np.max(spectrum[:, 1] + spectrum[:, 2])])
	plt.ticklabel_format(axis='y', style='scientific', scilimits=(0, 0))
	plt.xlabel("Energy after hohlraum wall (MeV)" if any_hohlraum else "Energy (MeV)")
	plt.ylabel("Yield (MeV⁻¹)")
	if tag != '':
		plt.title(f"{line_of_site}, position {position} ({tag})")
	elif position != "":
		plt.title(f"{line_of_site}, position {position}")
	else:
		plt.title(f"{shot_number}, {line_of_site}")
	plt_set_locators()
	plt.tight_layout()
	plt.savefig(filepath+'_spectrum.png', dpi=300)
	plt.savefig(filepath+'_spectrum.eps')
	if show_plots:
		plt.show()
	plt.close()

	# figure out the hohlraum correction for this LOS and position
	if '90' in line_of_site and any(parameters["hohlraum"].values()) > 0:
		if position in parameters["hohlraum"]:
			hohlraum_layers = parameters["hohlraum"][position]
		else:
			hohlraum_layers = parameters["hohlraum"][""]
	else:
		hohlraum_layers = []
	yeeld, mean, sigma = perform_hohlraum_correction(hohlraum_layers, (yeeld, mean, sigma))

	# do the ρR analysis for both the shock and compression peak
	try:
		rhoR = calculate_rhoR(mean, shot_day+shot_number, parameters)
	except ValueError as e:
		print(f"setting ρR to nan because {e}")
		rhoR = (nan, nan, nan)
	if good_compression_fit:
		compression_yield, compression_mean, compression_sigma = \
			perform_hohlraum_correction(hohlraum_layers,
			                            (compression_yield, compression_mean, compression_sigma))
		try:
			compression_rhoR = calculate_rhoR(
				compression_mean, shot_day+shot_number, parameters)
		except ValueError:
			compression_rhoR = (nan, nan, nan)
	else:
		compression_rhoR = (nan, nan, nan)

	# test_mean, _, _ = perform_correction(layers, 5, 0, 0)
	# test_rhoR, _, _, _ = calculate_rhoR(test_mean, 0, shot_day+shot_number, parameters)
	# print(f"\tthe maximum measurable ρR is {test_rhoR:.1f} mg/cm^2")

	return (
		shot_day, shot_number,
		line_of_site, position,
		tag, any_overlap_here, any_clipping_here,
		(yeeld,
		 mean,
		 sigma),
		rhoR,
		(compression_yield,
		 compression_mean,
		 compression_sigma),
		compression_rhoR,
		spectrum,
	)


def assign_label(item: np_Analysis, multiple_days: bool, multiple_shots: bool) -> tuple[str, str]:
	""" compose short human-readable strings that uniquely identify this shot and spectrum
	    :param item: the features of the spectrum being labelled
	    :param multiple_days: whether we need to specify which day we're talking about
	    :param multiple_shots: whether we need to specify which shot we're talking about
	    :return: a label for the shot and a label for the analysis
	"""
	if item["position"] == "":
		location = item["line_of_site"]
	else:
		location = f"{item['line_of_site']}:{item['position']}"
	if multiple_days and len(item["shot_number"]) < 4:
		shot_label = f"{item['shot_day']}-{item['shot_number']}"
	else:
		shot_label = item['shot_number']

	if multiple_shots:
		analysis_label = f"{shot_label}\n{location}"
	else:
		analysis_label = f"{location}"
	# add any non-empty tags as parentheticals
	if len(item["tag"]) > 0:
		analysis_label += f" ({item['tag']})"
	# put it on one line if it'll fit
	if len(analysis_label) < 13:
		analysis_label = analysis_label.replace('\n', ' ')

	return shot_label, analysis_label


def load_rhoR_parameters(folder: str) -> dict[str, Any]:
	""" load the analysis parameters given in shot_info.csv and hohlraum.txt file
	    :param folder: the absolute or relative path to the directory with the data we're considering
	    :return: a dictionary containing values for 'ablator thickness', 'hohlraum', and a bunch of other stuff.
	    :raise HohlraumFileError: if the hohlraum.txt file hasn't been created but this is a NIF shot
	"""
	# start by taking any relevant information from shot_info.csv
	params: dict[str, Any] = {}
	try:
		nif_shot_number = normalize_shot_number(os.path.basename(folder))
	except ValueError:
		nif_shot = False
	else:
		nif_shot = True
		nif_shot_table = pd.read_csv(
			"shot_info.csv", skipinitialspace=True, index_col="shot number", dtype={})
		shot_info = nif_shot_table.loc[nif_shot_number]
		for key in ["ablator radius", "ablator thickness", "ablator material",
		            "fill pressure", "deuterium fraction", "helium-3 fraction"]:
			if not pd.isnull(shot_info[key]):
				params[key] = shot_info[key]

		# calculate the converged shell thickness
		if "shell thickness" not in params:
			params["shell thickness"] = params["ablator thickness"]*40.0/200.0  # from "Alex's paper" (idk which)

	# read hohlraum.txt if it exists
	if not os.path.isfile(os.path.join(folder, "hohlraum.txt")):
		with open(os.path.join(folder, "hohlraum.txt"), "w"):
			pass  # create a blank file if there is none
	with open(os.path.join(folder, "hohlraum.txt"), encoding="utf-8") as f:
		hohlraum_codes = f.readlines()
	# NIF shots usually have hohlraeume, so complain if it looks like the user forgot to add it
	if nif_shot and len(hohlraum_codes) == 0:
		raise HohlraumFileError(f"you need to fill out `{os.path.join(os.getcwd(), folder, 'hohlraum.txt')}` with the "
		                        f"hohlraum information.  if there is no hohlraum, just put 'none'.")
	if len(hohlraum_codes) == 1 and hohlraum_codes[0].lower().strip() == "none":  # this is the explicit way to indicate no hohlraum
		hohlraum_codes = []

	# parse the hohlraum layers and booleans
	params["hohlraum"] = OrderedDict()
	params["clipping"] = set()
	params["overlap"] = set()
	for hohlraum_code in hohlraum_codes:
		# separate out all of the line of sight keys
		if ":" in hohlraum_code:
			key, layer_set_code = re.split(r"\s*:\s*", hohlraum_code, maxsplit=2)
		else:
			key, layer_set_code = "", hohlraum_code
		# check for clipped spectrum flags
		if "clip" in layer_set_code:
			params["clipping"].add(key)
			layer_set_code = re.sub(r"\s*\(?clip(ped|ping)?\)?\s*", "", layer_set_code)
		# check for track overlap flags
		if "overlap" in layer_set_code:
			params["overlap"].add(key)
			layer_set_code = re.sub(r"\s*\(?(track[ -]?)?overlap(ped|ping)?\)?\s*", "", layer_set_code)
		# parse the material thicknesses
		layers: list[Layer] = []
		for layer_code in re.split(r"\s+", layer_set_code):
			if len(layer_code) > 0:
				try:
					thickness, _, material = re.fullmatch(
						r"([0-9.]+)([uμ]m)?([A-Za-z0-9-]+)", layer_code).groups()
				except AttributeError:
					raise HohlraumFileError(f"I can't read the hohlraum code '{layer_code}'.")
				layers.append((float(thickness), material))
		params["hohlraum"][key] = layers

	return params


def gaussian(E, yeeld, mean, sigma):
	""" it’s just a gaussian. """
	return yeeld*np.exp(-(E - mean)**2/(2*sigma**2))/np.sqrt(2*pi*sigma**2)


def skew_gaussian(E_out, yeeld, median, birth_sigma, birth_energy=14.7):
	""" a spectrum that could result from a gaussian D3He peak getting ranged down thru some
	    material, approximately speaking (assume dE/dx \\propto E)
	"""
	valid = E_out**2 + (birth_energy**2 - median**2) > 0
	result = np.empty(E_out.shape)
	E_in = np.sqrt(E_out[valid]**2 + (birth_energy**2 - median**2))
	dEdE = E_out[valid]/E_in
	result[valid] = dEdE*gaussian(E_in, yeeld, birth_energy, birth_sigma)
	result[~valid] = 0
	return result


def fit_skew_gaussian(spectrum: NDArray[float], lower_bound: float, upper_bound: float) -> Peak:
	""" do a gaussian fit, accounting for how the shape changes and the peak shifts when the
	    spectrum has been ranged down to where σ !<< E (Fredrick and Alex don’t account for this in
	    their fitting, but I don’t think it’s significant above 7 MeV)
	"""
	spectrum = spectrum[(spectrum[:, 0] > lower_bound) & (spectrum[:, 0] < upper_bound), :]
	if spectrum.size == 0:
		raise FitError("the shock peak is too close to the edge for us to even try to see a compression peak")
	rainge = np.ptp(spectrum[:, 0])
	center = np.mean(spectrum[:, 0])

	try:
		values, covariances = optimize.curve_fit(
			skew_gaussian, xdata=spectrum[:, 0], ydata=spectrum[:, 1], sigma=spectrum[:, 2],
			p0=(abs(np.max(spectrum[:, 1]))*rainge, center, rainge/4),
			bounds=(0, inf), absolute_sigma=True)
	except (ValueError, RuntimeError) as e:
		raise FitError(f"the shock peak could not be fit because {e}")

	return ((values[0], sqrt(covariances[0, 0]), sqrt(covariances[0, 0])),
	        (values[1], sqrt(covariances[1, 1]), sqrt(covariances[1, 1])),
	        (values[2], sqrt(covariances[2, 2]), sqrt(covariances[2, 2])))


def plt_set_locators() -> None:
	try:
		plt.locator_params(steps=[1, 2, 5, 10])
	except TypeError:
		pass


def main():
	parser = argparse.ArgumentParser(
		prog="python make_plots_from_analysis.py",
		description = "take the analysis .csv files created by AnalyzeCR39 in a given series of folders, and "
		              "generate a bunch of plots and tables summarizing the information therein.")
	parser.add_argument(
		"folders", type=str,
		help="A list of subdirectories in data/ to search for analysis results. If you have multiple folders, separate "
		     "them with commas. If you're doing multiple shots from the same day, you don't have to include the day "
		     "part of the shot number after the first one (for example, 'N220420-001,-002,-003').")
	parser.add_argument(
		"--shell_material", type=str, default=None,
		help="The name of the capsule material, if it's not in shot_info.csv (for instance if it's an OMEGA shot). "
		     "Must be one of 'CH', 'CH2', 'HDC', 'SiO2', or 'Be'."
	)
	parser.add_argument(
		"--shell_density", type=float, default=None,
		help="The nominal shell density at bang-time, in g/mL. Only needed for OMEGA shots."
	)
	parser.add_argument(
		"--shell_temperature", type=float, default=None,
		help="The nominal electron temperature in the shell at bang-time, in keV. Only needed for OMEGA shots."
	)
	parser.add_argument(
		"--secondary", action="store_true",
		help="to treat the protons as secondary reactions (in which case the assumed mean birth energy is 15.0 MeV instead of 14.7)")
	parser.add_argument(
		"--show", action="store_true",
		help="to show the plots as they're generated in addition to saving them in the subdirectory."
	)
	args = parser.parse_args()

	options = dict()
	if args.shell_material is not None:
		options["ablator material"] = args.shell_material
	if args.shell_density is not None:
		options["shell density"] = args.shell_density
	if args.shell_temperature is not None:
		options["shell electron temperature"] = args.shell_temperature
	options["secondary"] = args.secondary

	make_plots_from_analysis(args.folders.split(","), args.show, options)


class FixedOrderFormatter(ScalarFormatter):
	"""Formats axis ticks using scientifick notacion with a constant ordre of magnitude"""
	def __init__(self, order_of_mag=0, use_offset=True, use_math_text=False):
		self._order_of_mag = order_of_mag
		ScalarFormatter.__init__(self, useOffset=use_offset,
		                         useMathText=use_math_text)
	def _set_orderOfMagnitude(self, _):
		"""Ovre-riding this to avoid having orderOfMagnitude reset elsewhere"""
		self.orderOfMagnitude = self._order_of_mag


class FitError(Exception):
	""" the error to throw when you can't find a good curve fit """
	pass


class HohlraumFileError(Exception):
	""" the error to throw when there's a problem with hohlraum.txt """
	pass


class MetadataNotFoundError(Exception):
	""" the error to throw when a file doesn't contain enuff information """
	pass


if __name__ == "__main__":
	main()
